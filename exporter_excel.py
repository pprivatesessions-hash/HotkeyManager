import logging
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from .models.analysis import AnalysisResult

logger = logging.getLogger(__name__)


def export_excel(result: AnalysisResult, output_path: str) -> str:
    logger.info(f"Экспорт Excel: {output_path}")

    wb = Workbook()
    ws = wb.active
    ws.title = "Горячие клавиши"

    headers = ["Категория", "Команда", "Текущая клавиша", "Новая клавиша", "Статус", "Комментарий"]
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

    status_colors = {
        "assigned": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
        "needs_assignment": PatternFill(
            start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"
        ),
        "duplicate": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
    }

    status_labels = {
        "assigned": "Назначена",
        "needs_assignment": "Требует назначения",
        "duplicate": "Дубликат",
    }

    for row_idx, cmd in enumerate(result.commands, 2):
        ws.cell(row=row_idx, column=1, value=cmd.category).border = thin_border
        ws.cell(row=row_idx, column=2, value=cmd.name).border = thin_border
        ws.cell(row=row_idx, column=3, value=cmd.current_hotkey or "—").border = thin_border
        ws.cell(row=row_idx, column=4, value=cmd.suggested_hotkey or "—").border = thin_border

        status_cell = ws.cell(
            row=row_idx, column=5, value=status_labels.get(cmd.status, cmd.status)
        )
        status_cell.border = thin_border
        if cmd.status in status_colors:
            status_cell.fill = status_colors[cmd.status]

        comment = cmd.conflict_with if cmd.conflict_with else ""
        ws.cell(row=row_idx, column=6, value=comment).border = thin_border

    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 40
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 22
    ws.column_dimensions["F"].width = 30

    ws.auto_filter.ref = f"A1:F{len(result.commands) + 1}"

    Path(output_path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)

    logger.info(f"Excel сохранён: {output_path}")
    return output_path
